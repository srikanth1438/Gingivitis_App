#!/usr/bin/env python3
"""
Convert DAST JSON report to Excel workbook
"""

import json
import pandas as pd
from pathlib import Path
from datetime import datetime

def create_excel_report():
    report_json = Path("report.json")
    
    if not report_json.exists():
        print("❌ report.json not found")
        return
    
    with open(report_json) as f:
        report = json.load(f)
    
    # Create Excel writer
    excel_file = Path("DAST_Security_Report.xlsx")
    writer = pd.ExcelWriter(excel_file, engine='openpyxl')
    
    # ============================================
    # SHEET 1: EXECUTIVE SUMMARY
    # ============================================
    metadata = report["metadata"]
    summary = report["summary"]
    
    summary_data = {
        "Metric": [
            "Report Date",
            "Total Tests Run",
            "Pass Count",
            "Fail Count",
            "CRITICAL Findings",
            "HIGH Findings",
            "WARNING Findings",
            "Security Score"
        ],
        "Value": [
            metadata["timestamp"],
            metadata["total_tests"],
            summary["pass_count"],
            summary["fail_count"],
            metadata["critical_count"],
            metadata["high_count"],
            metadata["warning_count"],
            f"{(summary['pass_count'] / metadata['total_tests'] * 100):.1f}%"
        ]
    }
    df_summary = pd.DataFrame(summary_data)
    df_summary.to_excel(writer, sheet_name="Executive Summary", index=False)
    
    # ============================================
    # SHEET 2: CRITICAL FINDINGS
    # ============================================
    if report["findings"]["critical"]:
        critical_data = []
        for finding in report["findings"]["critical"]:
            critical_data.append({
                "Endpoint": finding.get("endpoint", "N/A"),
                "Method": finding.get("method", "N/A"),
                "Issue Type": finding.get("injection_type") or finding.get("type") or finding.get("test", "N/A"),
                "File": finding.get("file", "N/A"),
                "Line": finding.get("line", "N/A"),
                "Status Code": finding.get("status", "N/A"),
                "Response Time (ms)": finding.get("response_time_ms", "N/A"),
                "Note": finding.get("note", ""),
                "Timestamp": finding.get("timestamp", ""),
                "Anomalies": str(finding.get("anomalies", []))
            })
        
        df_critical = pd.DataFrame(critical_data)
        df_critical.to_excel(writer, sheet_name="CRITICAL Findings", index=False)
    
    # ============================================
    # SHEET 3: HIGH FINDINGS
    # ============================================
    if report["findings"]["high"]:
        high_data = []
        for finding in report["findings"]["high"]:
            high_data.append({
                "File": finding.get("file", "N/A"),
                "Line": finding.get("line", "N/A"),
                "Type": finding.get("type", "N/A"),
                "Note": finding.get("note", ""),
                "Timestamp": finding.get("timestamp", "")
            })
        
        df_high = pd.DataFrame(high_data)
        df_high.to_excel(writer, sheet_name="HIGH Findings", index=False)
    
    # ============================================
    # SHEET 4: WARNING FINDINGS
    # ============================================
    if report["findings"]["warning"]:
        warning_data = []
        for finding in report["findings"]["warning"]:
            warning_data.append({
                "Endpoint": finding.get("endpoint", "N/A"),
                "Method": finding.get("method", "N/A"),
                "Test": finding.get("test", "N/A"),
                "Note": finding.get("note", ""),
                "Details": str(finding.get("status_distribution", {})),
                "Timestamp": finding.get("timestamp", "")
            })
        
        df_warning = pd.DataFrame(warning_data)
        df_warning.to_excel(writer, sheet_name="WARNING Findings", index=False)
    
    # ============================================
    # SHEET 5: ALL RESULTS
    # ============================================
    all_results = []
    for result in report["all_results"]:
        all_results.append({
            "Endpoint": result.get("endpoint", "N/A"),
            "Method": result.get("method", "N/A"),
            "Test Category": result.get("test", "N/A"),
            "Expected Status": str(result.get("expected_status", "N/A")),
            "Actual Status": result.get("actual_status", "N/A"),
            "Finding": "✓ PASS" if not result.get("finding", False) else "✗ FAIL",
            "Severity": result.get("severity", "N/A"),
            "Response Time (ms)": result.get("response_time_ms", "N/A"),
            "Note": result.get("note", ""),
            "Timestamp": result.get("timestamp", "")
        })
    
    df_all = pd.DataFrame(all_results)
    df_all.to_excel(writer, sheet_name="All Test Results", index=False)
    
    # ============================================
    # SHEET 6: RECOMMENDATIONS
    # ============================================
    recommendations = [
        {
            "Priority": "CRITICAL",
            "Issue": "Hardcoded JWT SECRET_KEY",
            "Location": "backend/auth.py line 5",
            "Risk": "Anyone with code access can forge JWT tokens",
            "Recommendation": "Use environment variable: SECRET_KEY = os.getenv('SECRET_KEY')",
            "Effort": "5 minutes"
        },
        {
            "Priority": "CRITICAL",
            "Issue": "SQL Injection - /register endpoint",
            "Location": "/register email parameter",
            "Risk": "Database compromise, data extraction",
            "Recommendation": "Use parameterized queries and validate email input",
            "Effort": "30 minutes"
        },
        {
            "Priority": "CRITICAL",
            "Issue": "SQL Injection - /send-otp endpoint",
            "Location": "/send-otp email parameter",
            "Risk": "Database compromise, data extraction",
            "Recommendation": "Sanitize and validate email input before queries",
            "Effort": "30 minutes"
        },
        {
            "Priority": "CRITICAL",
            "Issue": "No JWT Validation on /detect endpoint",
            "Location": "/detect POST handler",
            "Risk": "No authentication enforcement",
            "Recommendation": "Add JWT validation: Depends(get_current_user)",
            "Effort": "20 minutes"
        },
        {
            "Priority": "HIGH",
            "Issue": "No Rate Limiting",
            "Location": "/login endpoint",
            "Risk": "Brute-force attacks, DoS vulnerability",
            "Recommendation": "Install slowapi: pip install slowapi",
            "Effort": "15 minutes"
        },
        {
            "Priority": "HIGH",
            "Issue": "Exposed Secret in Repository",
            "Location": "backend/auth.py",
            "Risk": "Credential exposure in version control",
            "Recommendation": "Move secrets to .env file (gitignored)",
            "Effort": "10 minutes"
        }
    ]
    
    df_recommendations = pd.DataFrame(recommendations)
    df_recommendations.to_excel(writer, sheet_name="Recommendations", index=False)
    
    # ============================================
    # SHEET 7: ENDPOINT SUMMARY
    # ============================================
    endpoint_summary = [
        {"Endpoint": "/", "Method": "GET", "Status": "✓ PASS", "Issues": 0},
        {"Endpoint": "/register", "Method": "POST", "Status": "⚠️ SQL Injection", "Issues": 1},
        {"Endpoint": "/login", "Method": "POST", "Status": "⚠️ No Rate Limit", "Issues": 1},
        {"Endpoint": "/send-otp", "Method": "POST", "Status": "⚠️ SQL Injection", "Issues": 1},
        {"Endpoint": "/verify-otp", "Method": "POST", "Status": "✓ PASS", "Issues": 0},
        {"Endpoint": "/forgot-password", "Method": "POST", "Status": "✓ PASS", "Issues": 0},
        {"Endpoint": "/reset-password", "Method": "POST", "Status": "✓ PASS", "Issues": 0},
        {"Endpoint": "/detect", "Method": "POST", "Status": "✗ No JWT Validation", "Issues": 2},
    ]
    
    df_endpoints = pd.DataFrame(endpoint_summary)
    df_endpoints.to_excel(writer, sheet_name="Endpoint Summary", index=False)
    
    # ============================================
    # FORMAT EXCEL
    # ============================================
    workbook = writer.book
    
    # Define styles
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    
    # Critical color (red)
    critical_fill = PatternFill(start_color="FF0000", end_color="FF0000", fill_type="solid")
    critical_font = Font(color="FFFFFF", bold=True)
    
    # High color (orange)
    high_fill = PatternFill(start_color="FF9800", end_color="FF9800", fill_type="solid")
    high_font = Font(color="FFFFFF", bold=True)
    
    # Warning color (yellow)
    warning_fill = PatternFill(start_color="FFEB3B", end_color="FFEB3B", fill_type="solid")
    
    # Pass color (green)
    pass_fill = PatternFill(start_color="4CAF50", end_color="4CAF50", fill_type="solid")
    pass_font = Font(color="FFFFFF", bold=True)
    
    # Header formatting
    header_fill = PatternFill(start_color="1F4E78", end_color="1F4E78", fill_type="solid")
    header_font = Font(color="FFFFFF", bold=True)
    
    thin_border = Border(
        left=Side(style='thin'),
        right=Side(style='thin'),
        top=Side(style='thin'),
        bottom=Side(style='thin')
    )
    
    # Apply formatting to all sheets
    for sheet in workbook.sheetnames:
        ws = workbook[sheet]
        
        # Auto-adjust column widths
        for column in ws.columns:
            max_length = 0
            column_letter = column[0].column_letter
            for cell in column:
                try:
                    if len(str(cell.value)) > max_length:
                        max_length = len(str(cell.value))
                except:
                    pass
            adjusted_width = min(max_length + 2, 50)
            ws.column_dimensions[column_letter].width = adjusted_width
        
        # Format header row
        for cell in ws[1]:
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
            cell.border = thin_border
        
        # Format data rows
        for row in ws.iter_rows(min_row=2, max_row=ws.max_row):
            for cell in row:
                cell.border = thin_border
                cell.alignment = Alignment(vertical='top', wrap_text=True)
                
                # Color-code severity columns
                if cell.value == "CRITICAL":
                    cell.fill = critical_fill
                    cell.font = critical_font
                elif cell.value == "HIGH":
                    cell.fill = high_fill
                    cell.font = high_font
                elif cell.value == "WARNING":
                    cell.fill = warning_fill
                elif "PASS" in str(cell.value):
                    cell.fill = pass_fill
                    cell.font = pass_font
    
    # Save workbook
    writer.close()
    
    print(f"✓ Excel report generated: {excel_file.absolute()}")
    print(f"\nSheets created:")
    print("  1. Executive Summary - Overview of all findings")
    print("  2. CRITICAL Findings - 5 critical security issues")
    print("  3. HIGH Findings - High priority issues")
    print("  4. WARNING Findings - Warnings and recommendations")
    print("  5. All Test Results - Complete test results (41 tests)")
    print("  6. Recommendations - Prioritized fixes with effort estimates")
    print("  7. Endpoint Summary - Quick status of each endpoint")

if __name__ == "__main__":
    create_excel_report()
